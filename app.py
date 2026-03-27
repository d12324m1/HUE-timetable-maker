import streamlit as st
import pandas as pd
import pdfplumber
import unicodedata
import re
import io
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import time

# ==========================================
# 🧠 1. PDF解析・抽出ロジック（心臓部）
# ==========================================
def normalize_text(text):
    if pd.isna(text): return ""
    return unicodedata.normalize('NFKC', str(text)).replace(" ", "").replace("　", "").replace("\n", "").upper()

def extract_base_name(text):
    base = re.sub(r'[*＊\s]*[A-ZＡ-Ｚ0-9０-９Ⅰ-Ⅲ]+$', '', text)
    return re.sub(r'\(.*?\)$|（.*?）$', '', base)

def build_master_database(uploaded_file):
    all_data = []
    with pdfplumber.open(uploaded_file) as pdf:
        target_pages = pdf.pages[9:42] 
        total_pages = len(target_pages)
        progress_bar = st.progress(0, text="（1/2）開設科目一覧を解析中...☕")
        
        for i, page in enumerate(target_pages, start=10):
            tables = page.extract_tables()
            if not tables: continue
            for table in tables:
                df = pd.DataFrame(table).T
                df = df.iloc[:, ::-1].reset_index(drop=True)
                df = df.apply(lambda col: col.map(lambda x: str(x)[::-1] if pd.notna(x) else x))
                while len(df) > 0:
                    first_row_str = "".join(df.iloc[0].fillna('').astype(str))
                    if any(kw in first_row_str for kw in ["科目番号", "授業科目名", "番番号号", "単単位位"]):
                        df = df.iloc[1:].reset_index(drop=True)
                    else: break 
                df = df.ffill()
                try:
                    ext_df = df.iloc[:, [1, 3, 4, 5, 6, 7, 9]].copy()
                    ext_df.columns = ['科目番号', '授業科目名', '単位', '開講期', '曜日', '時限(講目)', '対象学年']
                    all_data.append(ext_df)
                except Exception: pass
            progress_bar.progress(int(((i - 9) / total_pages) * 100), text=f"（1/2）開設科目一覧を解析中... ({int(((i - 9) / total_pages) * 100)}%)")
        progress_bar.progress(100, text="（1/2）マスターデータ構築完了！✨")
        time.sleep(1)
        progress_bar.empty()

    if not all_data: return None
    master_df = pd.concat(all_data, ignore_index=True).replace('\n', '', regex=True)
    master_df['検索用科目名'] = master_df['授業科目名'].apply(normalize_text)
    master_df['ベース科目名'] = master_df['検索用科目名'].apply(extract_base_name)
    return master_df

def cross_reference(master_df, uploaded_req):
    req_text = ""
    with pdfplumber.open(uploaded_req) as pdf:
        num_pages = len(pdf.pages)
        progress_bar = st.progress(0, text="（2/2）履修基準とクロスチェック中...🔍")
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text: req_text += text
            progress_bar.progress(int(((i + 1) / num_pages) * 100), text=f"（2/2）履修基準とクロスチェック中... ({int(((i + 1) / num_pages) * 100)}%)")
        progress_bar.progress(100, text="（2/2）照合完了！🎉")
        time.sleep(1)
        progress_bar.empty()
                
    req_norm = normalize_text(req_text)
    hit_indices = [idx for idx, row in master_df.iterrows() if len(row['ベース科目名']) >= 3 and row['ベース科目名'] in req_norm]
    
    if hit_indices:
        return master_df.loc[hit_indices].reset_index(drop=True).drop(columns=['検索用科目名', 'ベース科目名'])
    return None

# ==========================================
# 🎨 2. Excel生成ロジック（手足）
# ==========================================
def generate_excel_bytes(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4年間履修計画"
    ws_hidden = wb.create_sheet("HiddenData")
    ws_hidden.sheet_state = 'hidden'
    hidden_row_counter = 1

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    spacer_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    input_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)

    schedule_dict = {}
    days_map = {"月": 0, "火": 1, "水": 2, "木": 3, "金": 4}
    
    for _, row in df.iterrows():
        subj, credit, term_str, day_str, period_str, grade_str = map(str, [row.get('授業科目名',''), row.get('単位','0'), row.get('開講期',''), row.get('曜日',''), row.get('時限(講目)',''), row.get('対象学年','1')])
        
        grade_match = re.search(r'\d+', grade_str)
        grade = min(int(grade_match.group()), 4) if grade_match else 1
        term_idx = 0 if "前" in term_str else 1
        
        day_idx = next((v for k, v in days_map.items() if k in day_str), -1)
        p_match = re.search(r'[1-5]', period_str)
        period_idx = int(p_match.group()) - 1 if p_match else -1
            
        if "集中" in day_str or "集中" in period_str or "不定期" in day_str or day_idx == -1 or period_idx == -1:
            day_idx, period_idx = -1, 5
            
        key = (grade, term_idx, day_idx, period_idx)
        schedule_dict.setdefault(key, []).append(f"{subj} [{credit}]")

    days = ["月曜日", "火曜日", "水曜日", "木曜日", "金曜日"]
    periods = ["1", "2", "3", "4", "5", "集中/不定期"]

    for grade in range(1, 5):
        start_row = (grade - 1) * 11 + 1
        for is_late, col_offset in enumerate([0, 7]):
            ws.cell(row=start_row, column=1+col_offset, value=f"{grade}年{'前期' if is_late==0 else '後期'}").fill = header_fill
            for i, day in enumerate(days):
                ws.cell(row=start_row, column=2+i+col_offset, value=day).fill = header_fill

            for p_idx, period in enumerate(periods):
                r = start_row + 1 + p_idx
                ws.cell(row=r, column=1+col_offset, value=period).fill = summary_fill if "集中" in period else header_fill
                ws.row_dimensions[r].height = 45
                
                for c_idx in range(5):
                    cell = ws.cell(row=r, column=2+c_idx+col_offset)
                    cell.border, cell.alignment, cell.fill = thin_border, center_align, input_fill
                    key = (grade, is_late, -1 if "集中" in period else c_idx, 5 if "集中" in period else p_idx)

                    if key in schedule_dict:
                        items = list(set(schedule_dict[key]))
                        for h_idx, item in enumerate(items, start=1):
                            ws_hidden.cell(row=hidden_row_counter, column=h_idx, value=item)
                        dv = DataValidation(type="list", formula1=f"HiddenData!$A${hidden_row_counter}:$Z${hidden_row_counter}", allow_blank=True)
                        ws.add_data_validation(dv)
                        dv.add(cell)
                        cell.value = f"▽ 選択 ({len(items)}件)"
                        hidden_row_counter += 1
                    else:
                        cell.fill = input_fill

            r_sum = start_row + 7
            ws.cell(row=r_sum, column=1+col_offset, value="予定単位数").fill = summary_fill
            formula_parts = [f"IFERROR(VALUE(MID({openpyxl.utils.get_column_letter(2+c+col_offset)}{r},FIND(\"[\",{openpyxl.utils.get_column_letter(2+c+col_offset)}{r})+1,FIND(\"]\",{openpyxl.utils.get_column_letter(2+c+col_offset)}{r})-FIND(\"[\",{openpyxl.utils.get_column_letter(2+c+col_offset)}{r})-1)),0)" for c in range(5) for r in range(start_row + 1, start_row + 7)]
            cell_sum_val = ws.cell(row=r_sum, column=2+col_offset, value="=" + "+".join(formula_parts))
            cell_sum_val.font = bold_font
            ws.merge_cells(start_row=r_sum, start_column=3+col_offset, end_row=r_sum, end_column=6+col_offset)

        if grade < 4:
            spacer_row = start_row + 9
            ws.row_dimensions[spacer_row].height = 12
            for col in range(1, 14):
                ws.cell(row=spacer_row, column=col).fill = spacer_fill

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None: cell.border, cell.alignment = thin_border, center_align

    ws.column_dimensions['A'].width = ws.column_dimensions['H'].width = 12
    ws.column_dimensions['G'].width = 3
    for col_letter in ['B','C','D','E','F', 'I','J','K','L','M']: ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# 🖥️ 3. Webアプリ UI（自動分岐搭載！）
# ==========================================
st.set_page_config(page_title="履修パズルメーカー", layout="centered")

st.title("北海道教育大学　時間割メーカー")
st.markdown("PDFをアップロードして、単位自動計算付きの**自分専用Excel時間割**を作成します。")
st.markdown("---")
st.warning("⚠️ **注意事項**\n仕様上、クロスチェック版には「一般教養科目」が含まれません。まずは専門科目を配置し、空きコマは「全科目カタログ版」を見ながら埋めるのがおすすめです！")

col1, col2 = st.columns(2)
with col1:
    st.subheader("1. 必須ファイル")
    uploaded_syllabus = st.file_uploader("『開設科目一覧』PDF", type="pdf", help="これをアップロードすると全科目が載った「カタログ版」が作れます。")

with col2:
    st.subheader("2. 任意ファイル（絞り込み）")
    uploaded_req = st.file_uploader("『専攻履修基準』PDF", type="pdf", help="追加でアップロードすると、必要な科目だけを抽出した「クロスチェック版」になります。")

st.markdown("---")

if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
    st.session_state.current_mode = None

if uploaded_syllabus:
    # 🌟 アップロード状況によってボタンの名前とモードを自動切り替え！
    mode_name = "履修基準クロスチェック版" if uploaded_req else "全科目カタログ版"
    btn_color = "primary" if uploaded_req else "secondary"
    
    if st.button(f"🚀 【{mode_name}】を作成する", type=btn_color, use_container_width=True):
        st.session_state.excel_data = None # 一旦リセット
        
        master_df = build_master_database(uploaded_syllabus)
        if master_df is not None:
            # クロスチェック版なら照合、カタログ版ならそのまま流す
            final_df = cross_reference(master_df, uploaded_req) if uploaded_req else master_df
            
            if final_df is not None:
                with st.spinner("Excelパズル盤を組み立てています...🔨"):
                    st.session_state.excel_data = generate_excel_bytes(final_df)
                    st.session_state.current_mode = mode_name
                    st.balloons()
            else:
                st.error("⚠️ 履修基準に一致する科目が見つかりませんでした。")
        else:
            st.error("⚠️ 科目一覧PDFからデータを抽出できませんでした。")

# ダウンロードボタン
if st.session_state.excel_data is not None:
    st.success(f"🎉 あなた専用の【{st.session_state.current_mode}】パズル盤が完成しました！")
    
    filename = "CrossCheck_Timetable.xlsx" if "クロスチェック" in st.session_state.current_mode else "Catalog_Timetable.xlsx"
    
    st.download_button(
        label=f"📥 {filename} をダウンロード",
        data=st.session_state.excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
